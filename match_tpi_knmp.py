"""
match_tpi_knmp.py — Match data harga TPI (SIHI/PIPP) ke lokasi KNMP

Langkah:
  1. Jalankan scrape_sihi.py untuk ambil data TPI aktual → data/tpi_harga.csv
  2. Load Lokasi KNMP.xlsx → daftar KNMP beserta kabupaten/provinsi
  3. Match TPI ke KNMP berdasarkan kabupaten (exact/fuzzy) atau provinsi (estimasi)
  4. Simpan ke data/knmp_tpi_match.csv

Kolom output: knmp_id, nama_knmp, kabupaten, provinsi, nama_tpi,
              komoditas, harga_aktual, tanggal, match_type

Penggunaan:
  python match_tpi_knmp.py
  python match_tpi_knmp.py --skip-scrape      # pakai tpi_harga.csv yang sudah ada
  python match_tpi_knmp.py --tanggal 2024-05-01
  python match_tpi_knmp.py --verbose
"""

import argparse
import csv
import difflib
import logging
import subprocess
import sys
from pathlib import Path

import openpyxl

# ── Konfigurasi ───────────────────────────────────────────────────────────────

KNMP_XLSX   = Path("Lokasi KNMP.xlsx")
TPI_CSV     = Path("data/tpi_harga.csv")
OUTPUT_CSV  = Path("data/knmp_tpi_match.csv")

FUZZY_THRESHOLD = 0.80   # SequenceMatcher ratio minimum untuk "fuzzy"

CSV_FIELDNAMES = [
    "knmp_id", "nama_knmp", "kabupaten", "provinsi",
    "nama_tpi", "komoditas", "harga_aktual", "tanggal", "match_type",
]

# ── Normalisasi nama wilayah ──────────────────────────────────────────────────

_STRIP_PREFIXES = (
    "kabupaten ", "kab. ", "kab ", "kota ", "kepulauan ", "kep. ", "kep ",
)

def normalize(s: str) -> str:
    """Lowercase, hapus prefix administratif, kolaps spasi."""
    if not s:
        return ""
    s = s.lower().strip()
    for prefix in _STRIP_PREFIXES:
        if s.startswith(prefix):
            s = s[len(prefix):]
            break
    return " ".join(s.split())


# ── Load KNMP dari Excel ──────────────────────────────────────────────────────

def load_knmp(path: Path) -> list[dict]:
    """Gabungkan sheet Survey + Operasional, deduplikasi by ID."""
    if not path.exists():
        logging.error("File KNMP tidak ditemukan: %s", path)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    seen_ids: set[int] = set()
    records: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1]:          # kolom ID kosong
                continue
            try:
                knmp_id = int(row[1])
            except (ValueError, TypeError):
                continue
            if knmp_id in seen_ids:
                continue
            seen_ids.add(knmp_id)
            records.append({
                "knmp_id":   knmp_id,
                "nama_knmp": str(row[2] or "").strip(),
                "kabupaten": str(row[5] or "").strip(),
                "provinsi":  str(row[6] or "").strip(),
            })

    return sorted(records, key=lambda r: r["knmp_id"])


# ── Load TPI CSV ──────────────────────────────────────────────────────────────

def load_tpi(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


# ── Jalankan scraper ──────────────────────────────────────────────────────────

def run_scrape(tanggal: str | None, verbose: bool) -> bool:
    cmd = [sys.executable, "scrape_sihi.py"]
    if tanggal:
        cmd += ["--tanggal", tanggal]
    if verbose:
        cmd += ["--verbose"]
    logging.info("Menjalankan: %s", " ".join(cmd))
    result = subprocess.run(cmd, capture_output=not verbose)
    if result.returncode != 0 and not verbose:
        stderr = result.stderr.decode(errors="replace").strip()
        if stderr:
            logging.warning("Scraper stderr: %s", stderr[:500])
    return result.returncode == 0


# ── Indeks TPI ────────────────────────────────────────────────────────────────

def _build_index(tpi_rows: list[dict], field: str) -> dict[str, list[dict]]:
    idx: dict[str, list[dict]] = {}
    for row in tpi_rows:
        key = normalize(row.get(field, ""))
        if key:
            idx.setdefault(key, []).append(row)
    return idx


def _fuzzy_lookup(key: str, idx: dict[str, list[dict]]) -> tuple[str, list[dict]]:
    """Coba exact lalu fuzzy. Return (match_type, rows)."""
    if key in idx:
        return "exact", idx[key]
    matches = difflib.get_close_matches(key, idx.keys(), n=1, cutoff=FUZZY_THRESHOLD)
    if matches:
        return "fuzzy", idx[matches[0]]
    return "", []


def _latest_by_komoditas(rows: list[dict]) -> dict[str, dict]:
    """Ambil satu row terbaru (tanggal terbesar) per komoditas."""
    best: dict[str, dict] = {}
    for row in rows:
        k = row.get("komoditas", "") or ""
        tgl = row.get("tanggal", "") or ""
        if k not in best or tgl > (best[k].get("tanggal") or ""):
            best[k] = row
    return best


# ── Matching utama ────────────────────────────────────────────────────────────

def match_all(knmp_list: list[dict], tpi_rows: list[dict]) -> list[dict]:
    kab_idx  = _build_index(tpi_rows, "kabupaten")
    prov_idx = _build_index(tpi_rows, "provinsi")

    output: list[dict] = []

    for knmp in knmp_list:
        norm_kab  = normalize(knmp["kabupaten"])
        norm_prov = normalize(knmp["provinsi"])

        match_type, matched = _fuzzy_lookup(norm_kab, kab_idx)

        if not matched:
            _, matched = _fuzzy_lookup(norm_prov, prov_idx)
            if matched:
                match_type = "estimasi"

        if not matched:
            output.append({
                "knmp_id":      knmp["knmp_id"],
                "nama_knmp":    knmp["nama_knmp"],
                "kabupaten":    knmp["kabupaten"],
                "provinsi":     knmp["provinsi"],
                "nama_tpi":     "",
                "komoditas":    "",
                "harga_aktual": "",
                "tanggal":      "",
                "match_type":   "tidak_ada_data",
            })
            continue

        latest = _latest_by_komoditas(matched)
        for komo, row in sorted(latest.items()):
            output.append({
                "knmp_id":      knmp["knmp_id"],
                "nama_knmp":    knmp["nama_knmp"],
                "kabupaten":    knmp["kabupaten"],
                "provinsi":     knmp["provinsi"],
                "nama_tpi":     row.get("nama_tpi", ""),
                "komoditas":    komo,
                "harga_aktual": row.get("harga", ""),
                "tanggal":      row.get("tanggal", ""),
                "match_type":   match_type,
            })

    return output


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Match harga TPI ke lokasi KNMP")
    parser.add_argument("--skip-scrape", action="store_true",
                        help="Lewati scraping, gunakan tpi_harga.csv yang sudah ada")
    parser.add_argument("--tanggal", metavar="YYYY-MM-DD",
                        help="Tanggal data yang di-scrape")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%H:%M:%S",
    )

    # 1. Scrape TPI
    if not args.skip_scrape:
        ok = run_scrape(args.tanggal, args.verbose)
        if not ok:
            logging.warning("Scraper selesai dengan error — tetap lanjut dengan data yang ada")

    # 2. Load data
    tpi_rows  = load_tpi(TPI_CSV)
    knmp_list = load_knmp(KNMP_XLSX)

    logging.info("TPI rows  : %d (dari %s)", len(tpi_rows), TPI_CSV)
    logging.info("KNMP      : %d lokasi", len(knmp_list))

    if not tpi_rows:
        logging.warning("Tidak ada data TPI — output hanya daftar KNMP tanpa harga (match_type=tidak_ada_data)")

    # 3. Match
    output_rows = match_all(knmp_list, tpi_rows)

    # 4. Simpan
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDNAMES)
        writer.writeheader()
        writer.writerows(output_rows)

    # Ringkasan
    by_type: dict[str, int] = {}
    for r in output_rows:
        by_type[r["match_type"]] = by_type.get(r["match_type"], 0) + 1

    logging.info("Output    : %d rows → %s", len(output_rows), OUTPUT_CSV)
    for mtype, count in sorted(by_type.items()):
        logging.info("  %-20s %d", mtype, count)


if __name__ == "__main__":
    main()
