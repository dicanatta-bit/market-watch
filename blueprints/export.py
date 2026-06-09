"""Export blueprint — Excel + PDF"""
from io import BytesIO
from datetime import date
from flask import Blueprint, send_file, flash, redirect, url_for
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from models import KnmpLocation, KnmpLocationSnapshot, CommodityPrice, db
from auth import superadmin_required

export_bp = Blueprint("export", __name__, url_prefix="/export")


def _latest_snapshots():
    sub = (
        db.session.query(
            KnmpLocationSnapshot.id_lokasi,
            db.func.max(KnmpLocationSnapshot.snapshot_date).label("max_date"),
        )
        .group_by(KnmpLocationSnapshot.id_lokasi)
        .subquery()
    )
    return (
        KnmpLocationSnapshot.query.join(
            sub,
            db.and_(
                KnmpLocationSnapshot.id_lokasi == sub.c.id_lokasi,
                KnmpLocationSnapshot.snapshot_date == sub.c.max_date,
            ),
        ).all()
    )


@export_bp.route("/excel")
@superadmin_required
def excel():
    """Export KNMP locations + progress to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "KNMP Locations"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "ID", "Nama Kampung", "Provinsi", "Kabupaten", "Kecamatan",
        "Lat", "Lon", "Status KNMP", "Progress (%)", "Realisasi Fisik",
        "Realisasi Keuangan", "Penyedia", "Nelayan", "Kapal", "Tahun",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    locations = KnmpLocation.query.order_by(KnmpLocation.provinsi).all()
    snapshots = {s.id_lokasi: s for s in _latest_snapshots()}

    for row, loc in enumerate(locations, 2):
        snap = snapshots.get(loc.id_lokasi)
        data = [
            loc.id_lokasi, loc.nama_kampung, loc.provinsi, loc.kabupaten,
            loc.kecamatan, loc.lat, loc.lon, loc.status_knmp,
            snap.progress_kumulatif if snap else None,
            snap.realisasi_fisik if snap else None,
            snap.realisasi_keuangan if snap else None,
            loc.penyedia, loc.jumlah_nelayan, loc.jumlah_kapal, loc.tahun,
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"KNMP_Locations_{date.today()}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@export_bp.route("/pdf")
@superadmin_required
def pdf():
    """Export KNMP summary to PDF."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=10*mm,
                            bottomMargin=10*mm, leftMargin=5*mm, rightMargin=5*mm)
    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph(f"<b>Laporan KNMP — {date.today().strftime('%d %B %Y')}</b>", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 5*mm))

    header = ["No", "Nama Kampung", "Provinsi", "Status", "Progress (%)",
              "Fisik (%)", "Keu (%)", "Penyedia"]
    data = [header]

    locations = KnmpLocation.query.order_by(KnmpLocation.provinsi).limit(500).all()
    snapshots = {s.id_lokasi: s for s in _latest_snapshots()}

    for i, loc in enumerate(locations, 1):
        snap = snapshots.get(loc.id_lokasi)
        data.append([
            str(i),
            loc.nama_kampung or "",
            loc.provinsi or "",
            loc.status_knmp or "",
            f"{snap.progress_kumulatif:.1f}" if snap and snap.progress_kumulatif else "-",
            f"{snap.realisasi_fisik:.1f}" if snap and snap.realisasi_fisik else "-",
            f"{snap.realisasi_keuangan:.1f}" if snap and snap.realisasi_keuangan else "-",
            (loc.penyedia or "")[:30],
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B3A6B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f4f8")]),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)

    filename = f"Laporan_KNMP_{date.today()}.pdf"
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype="application/pdf")
