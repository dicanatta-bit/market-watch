"""
Gabungkan sheet "KNMP Survey" + "KNMP Operasional" dari Lokasi KNMP.xlsx
→ peta Leaflet interaktif dengan integrasi harga komoditas per wilayah.

Output: output/knmp_YYYYMMDD.html  dan  knmp.html (root)
"""
import openpyxl, json, os, re, sys
from datetime import datetime

try:
    from google.oauth2.service_account import Credentials
    import gspread
    GSPREAD = True
except ImportError:
    GSPREAD = False

# ── Konstanta ─────────────────────────────────────────────────────────────────

SPREADSHEET_ID   = "1qAn5AsxdL5CliEQltMuqN1hkAy6L-FIcMb1YqMFbUyw"
SHEET_HARGA_WIL  = "Harga per Wilayah"
CREDENTIALS_FILE = "credentials.json"
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# Faktor wilayah (sama dengan generate_harga_wilayah.py — untuk fallback statis)
WILAYAH_FAKTOR = {
    "Jawa-Bali":  1.00,
    "Sumatera":   0.95,
    "Kalimantan": 0.92,
    "Sulawesi":   0.90,
    "NTT-NTB":    0.88,
    "Maluku":     0.85,
    "Papua":      0.85,
}

# Mapping PROVINSI (uppercase) → WILAYAH
PROVINSI_WILAYAH = {
    # Sumatera
    "ACEH": "Sumatera", "NANGGROE ACEH DARUSSALAM": "Sumatera",
    "SUMATERA UTARA": "Sumatera", "SUMATRA UTARA": "Sumatera",
    "SUMATERA BARAT": "Sumatera", "SUMATRA BARAT": "Sumatera",
    "RIAU": "Sumatera", "KEPULAUAN RIAU": "Sumatera",
    "JAMBI": "Sumatera",
    "SUMATERA SELATAN": "Sumatera", "SUMATRA SELATAN": "Sumatera",
    "BENGKULU": "Sumatera", "LAMPUNG": "Sumatera",
    "KEPULAUAN BANGKA BELITUNG": "Sumatera", "BANGKA BELITUNG": "Sumatera",
    # Jawa-Bali
    "DKI JAKARTA": "Jawa-Bali", "JAKARTA": "Jawa-Bali",
    "JAWA BARAT": "Jawa-Bali", "JAWA TENGAH": "Jawa-Bali",
    "DI YOGYAKARTA": "Jawa-Bali", "DAERAH ISTIMEWA YOGYAKARTA": "Jawa-Bali",
    "JAWA TIMUR": "Jawa-Bali", "BANTEN": "Jawa-Bali", "BALI": "Jawa-Bali",
    # Kalimantan
    "KALIMANTAN BARAT": "Kalimantan", "KALIMANTAN TENGAH": "Kalimantan",
    "KALIMANTAN SELATAN": "Kalimantan", "KALIMANTAN TIMUR": "Kalimantan",
    "KALIMANTAN UTARA": "Kalimantan",
    # Sulawesi
    "SULAWESI UTARA": "Sulawesi", "SULAWESI TENGAH": "Sulawesi",
    "SULAWESI SELATAN": "Sulawesi", "SULAWESI TENGGARA": "Sulawesi",
    "GORONTALO": "Sulawesi", "SULAWESI BARAT": "Sulawesi",
    # NTT-NTB
    "NUSA TENGGARA BARAT": "NTT-NTB", "NTB": "NTT-NTB",
    "NUSA TENGGARA TIMUR": "NTT-NTB", "NTT": "NTT-NTB",
    # Maluku
    "MALUKU": "Maluku", "MALUKU UTARA": "Maluku",
    # Papua
    "PAPUA": "Papua", "PAPUA BARAT": "Papua",
    "PAPUA PEGUNUNGAN": "Papua", "PAPUA SELATAN": "Papua",
    "PAPUA TENGAH": "Papua", "PAPUA BARAT DAYA": "Papua",
    "IRIAN JAYA BARAT": "Papua",
}

# Nama pendek komoditas untuk tampilan popup
_SHORT_NAMES = {
    "Udang Vaname (Litopenaeus vannamei)":              "Udang Vaname",
    "Udang Windu (Penaeus monodon)":                    "Udang Windu",
    "Nila (Oreochromis niloticus)":                     "Nila",
    "Tuna Sirip Kuning / Yellowfin (Thunnus albacares)":"Tuna Yellowfin",
    "Tuna Cakalang (Katsuwonus pelamis)":               "Tuna Cakalang",
    "Kakap Merah (Lutjanus spp.)":                      "Kakap Merah",
    "Kerapu (Epinephelus spp.)":                        "Kerapu",
    "Rumput Laut (Eucheuma cottonii)":                  "Rumput Laut",
    "Rumput Laut ATC/SRC (E. cottonii processed)":      "RL ATC/SRC",
    "Lobster (Panulirus ornatus) / Mutiara":            "Lobster Mutiara",
    "Lobster (Panulirus homarus) / Pasir":              "Lobster Pasir",
    "Bandeng (Chanos chanos)":                          "Bandeng",
    "Cumi-cumi (Loligo spp.)":                          "Cumi-cumi",
    "Patin (Pangasianodon hypophthalmus)":              "Patin",
}

# 3 komoditas prioritas per wilayah (tampil di popup)
PRIORITY_PER_WILAYAH = {
    "Jawa-Bali":  [("Udang Vaname (Litopenaeus vannamei)", "Size 50"),
                   ("Udang Vaname (Litopenaeus vannamei)", "Size 60"),
                   ("Bandeng (Chanos chanos)", "250-500 g")],
    "Sumatera":   [("Udang Vaname (Litopenaeus vannamei)", "Size 50"),
                   ("Udang Windu (Penaeus monodon)", "Size 20"),
                   ("Tuna Cakalang (Katsuwonus pelamis)", "-")],
    "Kalimantan": [("Udang Vaname (Litopenaeus vannamei)", "Size 50"),
                   ("Udang Windu (Penaeus monodon)", "Size 20"),
                   ("Patin (Pangasianodon hypophthalmus)", "Utuh/Hidup")],
    "Sulawesi":   [("Udang Windu (Penaeus monodon)", "Size 20"),
                   ("Tuna Sirip Kuning / Yellowfin (Thunnus albacares)", "Sashimi grade"),
                   ("Rumput Laut (Eucheuma cottonii)", "Kering")],
    "NTT-NTB":    [("Rumput Laut (Eucheuma cottonii)", "Kering"),
                   ("Lobster (Panulirus ornatus) / Mutiara", ">200 g"),
                   ("Tuna Cakalang (Katsuwonus pelamis)", "-")],
    "Maluku":     [("Tuna Sirip Kuning / Yellowfin (Thunnus albacares)", "Sashimi grade"),
                   ("Tuna Cakalang (Katsuwonus pelamis)", "-"),
                   ("Lobster (Panulirus ornatus) / Mutiara", ">200 g")],
    "Papua":      [("Tuna Sirip Kuning / Yellowfin (Thunnus albacares)", "Sashimi grade"),
                   ("Udang Windu (Penaeus monodon)", "Size 20"),
                   ("Lobster (Panulirus ornatus) / Mutiara", ">200 g")],
}

# Harga nasional Jawa-Bali (fallback jika Google Sheets tidak tersedia)
_BASE_HARGA = {
    ("Udang Vaname (Litopenaeus vannamei)", "Size 50"):               ("60.000 – 65.000", "3,55 – 3,64"),
    ("Udang Vaname (Litopenaeus vannamei)", "Size 60"):               ("55.000 – 60.000", "3,55"),
    ("Udang Windu (Penaeus monodon)", "Size 20"):                     ("100.000 – 120.000", "8,00 – 10,00"),
    ("Nila (Oreochromis niloticus)", "300-500 g"):                    ("22.000 – 28.000", "3,00 – 4,00"),
    ("Tuna Sirip Kuning / Yellowfin (Thunnus albacares)", "Sashimi grade"): ("60.000 – 80.000", "5,00 – 8,00"),
    ("Tuna Cakalang (Katsuwonus pelamis)", "-"):                      ("15.000 – 25.000", "1,50 – 2,50"),
    ("Kakap Merah (Lutjanus spp.)", "-"):                             ("50.000 – 70.000", "5,00 – 8,00"),
    ("Kerapu (Epinephelus spp.)", "Hidup (>500 g)"):                  ("100.000 – 150.000", "8,00 – 12,00"),
    ("Rumput Laut (Eucheuma cottonii)", "Kering"):                    ("6.000 – 7.000", "0,40 – 0,50"),
    ("Lobster (Panulirus ornatus) / Mutiara", ">200 g"):              ("280.000 – 380.000", "18,00 – 22,00"),
    ("Bandeng (Chanos chanos)", "250-500 g"):                         ("20.000 – 28.000", "1,80 – 2,50"),
    ("Patin (Pangasianodon hypophthalmus)", "Utuh/Hidup"):            ("15.000 – 22.000", "—"),
    ("Cumi-cumi (Loligo spp.)", "-"):                                 ("35.000 – 50.000", "3,50 – 5,00"),
}


# ── Helper functions ──────────────────────────────────────────────────────────

def _s(v): return str(v).strip() if v is not None else ""
def _f(v, d=0.0):
    try: return float(v) if v is not None else d
    except: return d
def _i(v, d=0):
    try: return int(v) if v is not None else d
    except: return d


def _parse_range_idr(s):
    if not s or str(s).strip() in ("", "—", "-"):
        return []
    s = re.sub(r"[Rp\s]", "", str(s).strip())
    s = s.replace(".", "").replace(",", ".")
    parts = re.split(r"\s*[–\-]\s*", s)
    try:
        return [float(p.strip()) for p in parts if p.strip()]
    except ValueError:
        return []


def _apply_faktor_idr(harga_str, faktor):
    nums = _parse_range_idr(harga_str)
    if not nums:
        return "—"
    adj = sorted(n * faktor for n in nums)
    if len(adj) == 2:
        lo = f"{adj[0]:,.0f}".replace(",", ".")
        hi = f"{adj[1]:,.0f}".replace(",", ".")
        return f"{lo} – {hi}"
    return f"{adj[0]:,.0f}".replace(",", ".")


# ── Load harga wilayah dari Google Sheets ─────────────────────────────────────

def load_harga_wilayah(spreadsheet_id):
    """
    Coba baca sheet 'Harga per Wilayah' dari Google Sheets.
    Returns dict: {wilayah: [{"k": short_name, "s": size, "t": tambak_str}]}
    Falls back ke _BASE_HARGA × WILAYAH_FAKTOR jika tidak tersedia.
    """
    if not GSPREAD or not os.path.exists(CREDENTIALS_FILE):
        print("[KNMP Map] credentials.json tidak ada — pakai data harga statis.")
        return _build_static_harga()

    try:
        creds  = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
        client = gspread.Client(auth=creds)
        ss     = client.open_by_key(spreadsheet_id)
        ws     = ss.worksheet(SHEET_HARGA_WIL)
        data   = ws.get_all_values()
        print(f"[KNMP Map] Sheet '{SHEET_HARGA_WIL}' dibaca: {len(data) - 1} baris.")
        return _parse_harga_sheet(data)
    except gspread.exceptions.WorksheetNotFound:
        print(f"[KNMP Map] Sheet '{SHEET_HARGA_WIL}' belum ada — pakai data harga statis.")
    except Exception as e:
        print(f"[KNMP Map] Gagal membaca harga wilayah: {type(e).__name__} — pakai statis.")
    return _build_static_harga()


def _parse_harga_sheet(data):
    """
    Parse sheet 'Harga per Wilayah' (header row + data rows).
    Kolom: Tanggal(0), Komoditas(1), Size(2), Wilayah(3), Tambak(4), Ekspor(5), ...
    """
    result = {w: [] for w in WILAYAH_FAKTOR}
    seen   = {w: set() for w in WILAYAH_FAKTOR}  # deduplikasi per (komoditas, size)

    for row in data[1:]:
        if len(row) < 5:
            continue
        komoditas = row[1].strip()
        size      = row[2].strip()
        wilayah   = row[3].strip()
        tambak    = row[4].strip()

        if wilayah not in result:
            continue
        key = (komoditas, size)
        if key in seen[wilayah]:
            continue
        seen[wilayah].add(key)

        short = _SHORT_NAMES.get(komoditas, komoditas.split("(")[0].strip()[:20])
        result[wilayah].append({"k": short, "s": size, "t": tambak or "—"})

    # Urutkan sesuai PRIORITY_PER_WILAYAH, ambil max 5 per wilayah
    for wilayah, prio in PRIORITY_PER_WILAYAH.items():
        rows = result.get(wilayah, [])
        if not rows:
            continue
        rows_map = {r["k"] + "|" + r["s"]: r for r in rows}
        ordered  = []
        for komoditas, size in prio:
            short = _SHORT_NAMES.get(komoditas, komoditas.split("(")[0].strip()[:20])
            k = short + "|" + size
            if k in rows_map:
                ordered.append(rows_map[k])
        # Tambah sisa hingga 5 entri
        for r in rows:
            if len(ordered) >= 5:
                break
            if r not in ordered:
                ordered.append(r)
        result[wilayah] = ordered[:5]

    return result


def _build_static_harga():
    """Bangun harga wilayah dari BASE_HARGA × WILAYAH_FAKTOR sebagai fallback."""
    result = {}
    for wilayah, faktor in WILAYAH_FAKTOR.items():
        prio = PRIORITY_PER_WILAYAH.get(wilayah, [])
        rows = []
        for komoditas, size in prio:
            base = _BASE_HARGA.get((komoditas, size))
            if not base:
                continue
            tambak_est = _apply_faktor_idr(base[0], faktor)
            short      = _SHORT_NAMES.get(komoditas, komoditas.split("(")[0].strip()[:20])
            rows.append({"k": short, "s": size, "t": tambak_est})
        result[wilayah] = rows
    return result


# ── Baca XLSX ─────────────────────────────────────────────────────────────────

XLSX = "Lokasi KNMP.xlsx"
wb   = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
markers = []

# Survey: No=0,ID=1,Nama=2,Desa=3,Kec=4,Kab=5,Prov=6,Tahap=7,Kat=8,Lat=9,Lon=10,Luas=11,Pend=12
ws_sv = wb["KNMP Survey"]
for row in ws_sv.iter_rows(min_row=2, values_only=True):
    if row[0] is None or not _s(row[0]).isdigit(): continue
    lat, lon = row[9], row[10]
    if lat is None or lon is None: continue
    markers.append({
        "sumber": "survey", "id": _s(row[1]), "nama": _s(row[2]),
        "desa": _s(row[3]), "kecamatan": _s(row[4]), "kabupaten": _s(row[5]),
        "provinsi": _s(row[6]).upper(), "tahap": _s(row[7]), "kategori": _s(row[8]),
        "lat": float(lat), "lon": float(lon), "luas": _s(row[11]), "pendamping": _s(row[12]),
        "tahun": "", "penyedia": "", "nilai_p": 0, "pengawas": "",
        "realisasi": 0, "progres": None, "kondisi": "", "sarAda": 0, "sarNA": 0,
    })

# Operasional: No=0,ID=1,Nama=2,Desa=3,Kec=4,Kab=5,Prov=6,Tahap=7,Tahun=8,
#   Lat=9,Lon=10,Penyedia=11,NoKP=12,NilaiP=13,Pengawas=14,NoKW=15,NilaiW=16,
#   RealP=17,RealW=18,Progres=19,Kondisi=20,Kat=21,SarAda=22,SarNA=23
ws_op = wb["KNMP Operasional"]
for row in ws_op.iter_rows(min_row=2, values_only=True):
    if row[0] is None or not _s(row[0]).isdigit(): continue
    lat, lon = row[9], row[10]
    if lat is None or lon is None: continue
    markers.append({
        "sumber": "operasional", "id": _s(row[1]), "nama": _s(row[2]),
        "desa": _s(row[3]), "kecamatan": _s(row[4]), "kabupaten": _s(row[5]),
        "provinsi": _s(row[6]).upper(), "tahap": _s(row[7]), "kategori": _s(row[21]),
        "lat": float(lat), "lon": float(lon), "luas": "", "pendamping": "",
        "tahun": _s(row[8]), "penyedia": _s(row[11]), "nilai_p": _i(row[13]),
        "pengawas": _s(row[14]), "realisasi": _i(row[17]), "progres": _f(row[19]),
        "kondisi": _s(row[20]), "sarAda": _i(row[22]), "sarNA": _i(row[23]),
    })

wb.close()

# ── Stats ─────────────────────────────────────────────────────────────────────
total     = len(markers)
n_sv      = sum(1 for m in markers if m["sumber"] == "survey")
n_op      = sum(1 for m in markers if m["sumber"] == "operasional")
selesai   = sum(1 for m in markers if m["sumber"] == "operasional" and (m["progres"] or 0) >= 100)
tot_nilai = sum(m["nilai_p"] for m in markers)
prov_all  = sorted(set(m["provinsi"] for m in markers if m["provinsi"]))


def fmt_rp(n):
    if n >= 1e12: return f"Rp {n/1e12:.2f} T".replace(".", ",")
    if n >= 1e9:  return f"Rp {n/1e9:.2f} M".replace(".", ",")
    if n >= 1e6:  return f"Rp {n/1e6:.0f} Jt"
    return f"Rp {n:,.0f}"


# ── Load harga wilayah ────────────────────────────────────────────────────────
spreadsheet_id_arg = next(
    (a for a in sys.argv[1:] if not a.startswith("--")),
    SPREADSHEET_ID,
)
harga_wilayah = load_harga_wilayah(spreadsheet_id_arg)
tgl_harga     = datetime.now().strftime("%d/%m/%Y")

# Serialisasi ke JSON untuk embed di HTML
harga_wil_json  = json.dumps(harga_wilayah, ensure_ascii=False, separators=(",", ":"))
prov_wil_json   = json.dumps(PROVINSI_WILAYAH, ensure_ascii=False, separators=(",", ":"))

tgl      = datetime.now().strftime("%d %B %Y")
tgl_file = datetime.now().strftime("%Y%m%d")
mrk_json = json.dumps(markers, ensure_ascii=False, separators=(",", ":"))
prov_opts = "\n          ".join(
    f'<option value="{p}">{p.title()}</option>' for p in prov_all
)

# ── HTML ──────────────────────────────────────────────────────────────────────
html = f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Peta KNMP — PT Agrinas Jaladri Nusantara</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css">
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%;font-family:'Segoe UI',system-ui,sans-serif;background:#0d2244;overflow:hidden}}
#hdr{{position:fixed;top:0;left:0;right:0;z-index:1000;height:56px;background:linear-gradient(135deg,#1B3A6B 0%,#0d2244 100%);display:flex;align-items:center;justify-content:space-between;padding:0 20px;border-bottom:1px solid rgba(201,168,76,.35)}}
#hdr-brand{{display:flex;flex-direction:column;gap:1px}}
#hdr-title{{font-size:1.05rem;font-weight:700;color:#C9A84C;letter-spacing:.3px}}
#hdr-sub{{font-size:.68rem;color:rgba(203,213,225,.75)}}
#hdr-right{{display:flex;align-items:center;gap:10px}}
.hdr-chip{{font-size:.75rem;background:rgba(201,168,76,.15);border:1px solid rgba(201,168,76,.4);color:#C9A84C;padding:4px 11px;border-radius:20px;white-space:nowrap}}
#hdr-date{{font-size:.7rem;color:#64748b}}
#layout{{display:flex;height:100vh;padding-top:56px}}
#sidebar{{width:275px;flex-shrink:0;background:#162d55;overflow-y:auto;overflow-x:hidden;padding:12px;display:flex;flex-direction:column;gap:9px;border-right:1px solid rgba(201,168,76,.15)}}
#sidebar::-webkit-scrollbar{{width:3px}}
#sidebar::-webkit-scrollbar-thumb{{background:rgba(201,168,76,.35);border-radius:2px}}
.panel{{background:rgba(255,255,255,.05);border:1px solid rgba(201,168,76,.18);border-radius:8px;padding:10px 12px}}
.ptitle{{font-size:.63rem;font-weight:700;color:#C9A84C;text-transform:uppercase;letter-spacing:.9px;margin-bottom:8px}}
.stats-grid{{display:grid;grid-template-columns:1fr 1fr;gap:7px}}
.stat-box{{background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:6px;padding:7px 9px}}
.stat-num{{font-size:1.35rem;font-weight:800;color:#C9A84C;line-height:1}}
.stat-lbl{{font-size:.62rem;color:#94a3b8;margin-top:2px}}
.stat-box.wide{{grid-column:1/-1}}
.stat-num.sm{{font-size:.95rem}}
#search{{width:100%;padding:7px 10px;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.14);border-radius:6px;color:#fff;font-size:.82rem;outline:none}}
#search::placeholder{{color:rgba(255,255,255,.32)}}
#search:focus{{border-color:rgba(201,168,76,.55)}}
.flbl{{display:block;font-size:.7rem;color:#94a3b8;margin-bottom:3px}}
.fgap{{height:7px}}
select{{width:100%;padding:7px 10px;background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.14);border-radius:6px;color:#fff;font-size:.79rem;outline:none;cursor:pointer;-webkit-appearance:none}}
select option{{background:#1B3A6B}}
select:focus{{border-color:rgba(201,168,76,.55)}}
#btn-reset{{width:100%;padding:7px;margin-top:9px;background:rgba(201,168,76,.1);border:1px solid rgba(201,168,76,.38);border-radius:6px;color:#C9A84C;font-size:.8rem;font-weight:700;cursor:pointer;transition:background .15s;letter-spacing:.3px}}
#btn-reset:hover{{background:rgba(201,168,76,.25)}}
#filter-count{{font-size:.73rem;color:#60a5fa;margin-top:7px;display:none;text-align:center}}
.leg-item{{display:flex;align-items:center;gap:7px;padding:3px 0;font-size:.76rem;color:#cbd5e1}}
.leg-dot{{width:11px;height:11px;border-radius:50%;flex-shrink:0;border:1.5px solid rgba(255,255,255,.22)}}
.leg-star{{width:18px;height:18px;border-radius:50%;flex-shrink:0;border:1.5px solid #fff;display:flex;align-items:center;justify-content:center;font-size:10px;color:#fff}}
.leg-cnt{{margin-left:auto;font-size:.67rem;color:#94a3b8;font-variant-numeric:tabular-nums}}
.leg-sep{{font-size:.6rem;color:#C9A84C;font-weight:700;text-transform:uppercase;letter-spacing:.7px;margin:5px 0 2px}}
#map{{flex:1;z-index:0}}
.leaflet-popup-content-wrapper{{border-radius:9px;padding:0;box-shadow:0 6px 24px rgba(0,0,0,.35)}}
.leaflet-popup-content{{margin:0;min-width:240px;max-width:340px}}
.pu-head{{color:#C9A84C;font-weight:700;font-size:.82rem;padding:9px 13px;border-radius:9px 9px 0 0;line-height:1.35;background:linear-gradient(135deg,#1B3A6B,#0d2244)}}
.pu-tbl{{width:100%;border-collapse:collapse;font-size:.76rem}}
.pu-tbl td{{padding:4px 11px;vertical-align:top}}
.pu-tbl tr:nth-child(even){{background:#f1f5f9}}
.pu-tbl td:first-child{{font-weight:600;color:#475569;width:90px;white-space:nowrap}}
.pu-tbl td:last-child{{color:#1e293b;word-break:break-word}}
.pu-badge{{display:inline-block;padding:1px 8px;border-radius:10px;font-size:.7rem;font-weight:700;color:#fff}}
.pu-harga{{padding:6px 11px 8px;background:#f0f7ff;border-top:1px solid #dbeafe}}
.pu-harga-title{{font-size:.67rem;font-weight:700;color:#1B3A6B;text-transform:uppercase;letter-spacing:.5px;margin-bottom:5px}}
.pu-harga-row{{display:flex;justify-content:space-between;align-items:baseline;font-size:.71rem;padding:2px 0;border-bottom:1px dotted #e2e8f0}}
.pu-harga-k{{color:#475569;flex:1}}
.pu-harga-v{{color:#1B3A6B;font-weight:700;white-space:nowrap;margin-left:6px}}
.pu-harga-note{{font-size:.6rem;color:#94a3b8;margin-top:4px}}
@media(max-width:600px){{#sidebar{{display:none}}#hdr-date{{display:none}}}}
</style>
</head>
<body>

<div id="hdr">
  <div id="hdr-brand">
    <div id="hdr-title">Peta KNMP Nasional</div>
    <div id="hdr-sub">PT Agrinas Jaladri Nusantara (Persero)</div>
  </div>
  <div id="hdr-right">
    <div class="hdr-chip">{total:,} Lokasi</div>
    <div id="hdr-date">{tgl}</div>
  </div>
</div>

<div id="layout">
  <div id="sidebar">

    <div class="panel">
      <div class="ptitle">Ringkasan</div>
      <div class="stats-grid">
        <div class="stat-box"><div class="stat-num" id="stat-total">{total}</div><div class="stat-lbl">Total Lokasi</div></div>
        <div class="stat-box"><div class="stat-num" id="stat-selesai">{selesai}</div><div class="stat-lbl">Ops Selesai</div></div>
        <div class="stat-box"><div class="stat-num" id="stat-sv">{n_sv}</div><div class="stat-lbl">KNMP Survey</div></div>
        <div class="stat-box"><div class="stat-num" id="stat-op">{n_op}</div><div class="stat-lbl">KNMP Operasional</div></div>
        <div class="stat-box wide"><div class="stat-num sm">{fmt_rp(tot_nilai)}</div><div class="stat-lbl">Total Nilai Kontrak Penyedia</div></div>
      </div>
      <div id="filter-count"></div>
    </div>

    <div class="panel">
      <div class="ptitle">Cari KNMP</div>
      <input type="text" id="search" placeholder="Nama, kab, kec, penyedia…">
    </div>

    <div class="panel">
      <div class="ptitle">Filter</div>
      <label class="flbl">Provinsi</label>
      <select id="filter-prov">
        <option value="">Semua Provinsi</option>
          {prov_opts}
      </select>
      <div class="fgap"></div>
      <label class="flbl">Kategori KNMP</label>
      <select id="filter-kat">
        <option value="">Semua Kategori</option>
        <option value="2">Kategori 2 — Prioritas</option>
        <option value="1">Kategori 1 — Potensial</option>
        <option value="0">Kategori 0 — Observasi</option>
      </select>
      <div class="fgap"></div>
      <label class="flbl">Jenis Data</label>
      <select id="filter-sumber">
        <option value="">Semua Jenis</option>
        <option value="survey">Survey</option>
        <option value="operasional">Operasional ★</option>
      </select>
      <button id="btn-reset">&#x21BA; Reset Filter</button>
    </div>

    <div class="panel">
      <div class="ptitle">Legenda</div>
      <div class="leg-sep">Kategori</div>
      <div class="leg-item"><span class="leg-dot" style="background:#10B981"></span><span>Kategori 2 · Prioritas</span><span class="leg-cnt" id="lc2"></span></div>
      <div class="leg-item"><span class="leg-dot" style="background:#F59E0B"></span><span>Kategori 1 · Potensial</span><span class="leg-cnt" id="lc1"></span></div>
      <div class="leg-item"><span class="leg-dot" style="background:#6B7280"></span><span>Kategori 0 · Observasi</span><span class="leg-cnt" id="lc0"></span></div>
      <div class="leg-item"><span class="leg-dot" style="background:#94a3b8"></span><span>Tidak Terkategori</span><span class="leg-cnt" id="lcN"></span></div>
      <div class="leg-sep">Jenis</div>
      <div class="leg-item"><span class="leg-dot" style="background:#64748b;border-color:#999"></span><span>Survey (lingkaran)</span><span class="leg-cnt" id="lcSv"></span></div>
      <div class="leg-item"><span class="leg-star" style="background:#64748b;">★</span><span>Operasional (bintang)</span><span class="leg-cnt" id="lcOp"></span></div>
    </div>

  </div>

  <div id="map"></div>
</div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<script>
const DATA = {mrk_json};
const HARGA_WILAYAH   = {harga_wil_json};
const PROVINSI_WILAYAH = {prov_wil_json};
const HARGA_TGL = '{tgl_harga}';

const KAT_COLOR = {{'2':'#10B981','1':'#F59E0B','0':'#6B7280','':'#94a3b8'}};
const KAT_LABEL = {{'2':'Kat 2 · Prioritas','1':'Kat 1 · Potensial','0':'Kat 0 · Observasi','':'—'}};
const STS_COLOR = {{'selesai':'#10B981','berjalan':'#F59E0B','belum':'#EF4444'}};
const STS_LABEL = {{'selesai':'Selesai','berjalan':'Berjalan','belum':'Belum Mulai'}};

function esc(s){{return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')}}

function fmtRp(n){{
  if(!n||n===0)return'-';
  if(n>=1e12)return'Rp '+(n/1e12).toFixed(2).replace('.',',')+' T';
  if(n>=1e9)return'Rp '+(n/1e9).toFixed(2).replace('.',',')+' M';
  if(n>=1e6)return'Rp '+(n/1e6).toFixed(0)+' Jt';
  return'Rp '+n.toLocaleString('id');
}}

function statusOf(p){{
  if(p===null||p===undefined)return null;
  if(p>=100)return'selesai';
  if(p>0)return'berjalan';
  return'belum';
}}

function buildHargaHtml(provinsi){{
  const wil=PROVINSI_WILAYAH[provinsi]||null;
  if(!wil||!HARGA_WILAYAH[wil]||!HARGA_WILAYAH[wil].length)return'';
  const rows=HARGA_WILAYAH[wil].slice(0,3);
  const rowsHtml=rows.map(h=>
    `<div class="pu-harga-row"><span class="pu-harga-k">${{esc(h.k)}} <em style="color:#94a3b8;font-style:normal">${{esc(h.s)}}</em></span><span class="pu-harga-v">Rp ${{esc(h.t)}}/kg</span></div>`
  ).join('');
  return`<div class="pu-harga"><div class="pu-harga-title">&#128722; Harga Komoditas &mdash; ${{esc(wil)}}</div>${{rowsHtml}}<div class="pu-harga-note">Per ${{HARGA_TGL}} &middot; Estimasi harga di tingkat nelayan/tambak</div></div>`;
}}

const map = L.map('map',{{center:[-2.5,118],zoom:5}});
L.tileLayer('https://{{s}}.basemaps.cartocdn.com/rastertiles/voyager/{{z}}/{{x}}/{{y}}{{r}}.png',{{
  attribution:'&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a> &copy; <a href="https://carto.com/attributions">CARTO</a>',
  subdomains:'abcd',maxZoom:19
}}).addTo(map);

const cluster = L.markerClusterGroup({{
  chunkedLoading:true,
  maxClusterRadius:55,
  iconCreateFunction:function(c){{
    const n=c.getChildCount();
    const sz=n<10?30:n<30?36:n<70?42:48;
    return L.divIcon({{
      html:`<div style="background:#1B3A6B;border:2.5px solid #C9A84C;border-radius:50%;width:${{sz}}px;height:${{sz}}px;display:flex;align-items:center;justify-content:center;color:#C9A84C;font-size:${{sz<36?11:13}}px;font-weight:800;">${{n}}</div>`,
      className:'',iconSize:[sz,sz],iconAnchor:[sz/2,sz/2]
    }});
  }}
}});

const allMarkers = DATA.map(d=>{{
  const color = KAT_COLOR[d.kategori]||'#94a3b8';
  let m;
  const hargaHtml = buildHargaHtml(d.provinsi);

  if(d.sumber==='operasional'){{
    m = L.marker([d.lat,d.lon],{{
      icon:L.divIcon({{
        html:`<div style="width:20px;height:20px;border-radius:50%;background:${{color}};border:2px solid #fff;display:flex;align-items:center;justify-content:center;font-size:12px;line-height:1;color:#fff;box-shadow:0 2px 5px rgba(0,0,0,.45);">&#9733;</div>`,
        className:'',iconSize:[20,20],iconAnchor:[10,10],popupAnchor:[0,-10]
      }})
    }});
    const sts=statusOf(d.progres);
    const sc=sts?STS_COLOR[sts]:color;
    const pct=d.progres!=null?Math.round(d.progres):null;
    const progBar=pct!=null
      ?`<div style="height:5px;background:#e2e8f0"><div style="width:${{pct}}%;height:5px;background:${{sc}};border-radius:0 2px 2px 0"></div></div>`
      :'';
    const stsRow =sts?`<tr><td>Status</td><td><span class="pu-badge" style="background:${{sc}}">${{STS_LABEL[sts]}} ${{pct}}%</span></td></tr>`:'';
    const nilaiRow=d.nilai_p?`<tr><td>Kontrak</td><td>${{fmtRp(d.nilai_p)}}</td></tr>`:'';
    const rlsRow =d.realisasi?`<tr><td>Realisasi</td><td>${{fmtRp(d.realisasi)}}</td></tr>`:'';
    const kondRow=d.kondisi?`<tr><td>Kondisi</td><td>${{esc(d.kondisi.slice(0,120))}}</td></tr>`:'';
    m.bindPopup(`
      <div class="pu-head">&#9733; KNMP Operasional #${{esc(d.id)}} &mdash; ${{esc(d.nama)}}</div>
      ${{progBar}}
      <table class="pu-tbl">
        <tr><td>Tahap/Tahun</td><td>Tahap ${{esc(d.tahap)}} &middot; ${{esc(d.tahun)}}</td></tr>
        ${{stsRow}}
        <tr><td>Desa/Kel.</td><td>${{esc(d.desa)}}</td></tr>
        <tr><td>Kecamatan</td><td>${{esc(d.kecamatan)}}</td></tr>
        <tr><td>Kabupaten</td><td>${{esc(d.kabupaten)}}</td></tr>
        <tr><td>Provinsi</td><td>${{esc(d.provinsi)}}</td></tr>
        <tr><td>Penyedia</td><td>${{esc(d.penyedia)}}</td></tr>
        ${{nilaiRow}}${{rlsRow}}
        <tr><td>Pengawas</td><td>${{esc(d.pengawas)}}</td></tr>
        <tr><td>Sarpras</td><td>Ada: ${{d.sarAda}} &nbsp; N/A: ${{d.sarNA}}</td></tr>
        ${{kondRow}}
      </table>
      ${{hargaHtml}}
    `,{{maxWidth:340}});
  }} else {{
    m = L.circleMarker([d.lat,d.lon],{{
      radius:5,fillColor:color,color:'#fff',weight:1.5,opacity:1,fillOpacity:.85
    }});
    m.bindPopup(`
      <div class="pu-head">KNMP Survey #${{esc(d.id)}} &mdash; ${{esc(d.nama)}}</div>
      <table class="pu-tbl">
        <tr><td>Kategori</td><td><span class="pu-badge" style="background:${{color}}">${{KAT_LABEL[d.kategori]||'—'}}</span></td></tr>
        <tr><td>Tahap</td><td>${{esc(d.tahap)||'—'}}</td></tr>
        <tr><td>Desa/Kel.</td><td>${{esc(d.desa)}}</td></tr>
        <tr><td>Kecamatan</td><td>${{esc(d.kecamatan)}}</td></tr>
        <tr><td>Kabupaten</td><td>${{esc(d.kabupaten)}}</td></tr>
        <tr><td>Provinsi</td><td>${{esc(d.provinsi)}}</td></tr>
        <tr><td>Luas</td><td>${{esc(d.luas)||'—'}}</td></tr>
        <tr><td>Pendamping</td><td>${{esc(d.pendamping)||'—'}}</td></tr>
      </table>
      ${{hargaHtml}}
    `,{{maxWidth:340}});
  }}
  m._d = d;
  return m;
}});

cluster.addLayers(allMarkers);
map.addLayer(cluster);

function updLegend(arr){{
  const cc={{'2':0,'1':0,'0':0,'':0,sv:0,op:0}};
  arr.forEach(m=>{{
    const k=m._d.kategori;
    if(k==='2'||k==='1'||k==='0')cc[k]++;else cc['']++;
    if(m._d.sumber==='survey')cc.sv++;else cc.op++;
  }});
  document.getElementById('lc2').textContent=cc['2'];
  document.getElementById('lc1').textContent=cc['1'];
  document.getElementById('lc0').textContent=cc['0'];
  document.getElementById('lcN').textContent=cc[''];
  document.getElementById('lcSv').textContent=cc.sv;
  document.getElementById('lcOp').textContent=cc.op;
}}
updLegend(allMarkers);

const searchEl=document.getElementById('search');
const provEl  =document.getElementById('filter-prov');
const katEl   =document.getElementById('filter-kat');
const sbrEl   =document.getElementById('filter-sumber');
const btnReset=document.getElementById('btn-reset');
const fcountEl=document.getElementById('filter-count');
const stTotal =document.getElementById('stat-total');
const stSv    =document.getElementById('stat-sv');
const stOp    =document.getElementById('stat-op');
const stSel   =document.getElementById('stat-selesai');

function applyFilters(){{
  const q   =searchEl.value.toLowerCase().trim();
  const prov=provEl.value;
  const kat =katEl.value;
  const sbr =sbrEl.value;

  const filtered=allMarkers.filter(m=>{{
    const d=m._d;
    if(prov && d.provinsi!==prov) return false;
    if(kat  && d.kategori!==kat)  return false;
    if(sbr  && d.sumber!==sbr)    return false;
    if(q && ![d.nama,d.kabupaten,d.kecamatan,d.desa,d.penyedia,d.pendamping]
             .some(v=>v.toLowerCase().includes(q))) return false;
    return true;
  }});

  cluster.clearLayers();
  cluster.addLayers(filtered);

  const n  =filtered.length;
  const nOp=filtered.filter(m=>m._d.sumber==='operasional').length;
  stTotal.textContent=n;
  stSv.textContent   =n-nOp;
  stOp.textContent   =nOp;
  stSel.textContent  =filtered.filter(m=>m._d.sumber==='operasional'&&(m._d.progres||0)>=100).length;

  if(n<allMarkers.length){{
    fcountEl.style.display='block';
    fcountEl.textContent=`dari ${{allMarkers.length.toLocaleString('id')}} total`;
  }} else {{
    fcountEl.style.display='none';
  }}

  updLegend(filtered);

  if(n>0&&n<allMarkers.length){{
    const bnd=L.featureGroup(filtered).getBounds();
    if(bnd.isValid())map.fitBounds(bnd,{{padding:[40,40],maxZoom:13}});
  }}
}}

[searchEl,provEl,katEl,sbrEl].forEach(el=>el.addEventListener(
  el.tagName==='INPUT'?'input':'change', applyFilters
));

btnReset.addEventListener('click',()=>{{
  searchEl.value='';provEl.value='';katEl.value='';sbrEl.value='';
  stTotal.textContent={total};
  stSv.textContent   ={n_sv};
  stOp.textContent   ={n_op};
  stSel.textContent  ={selesai};
  fcountEl.style.display='none';
  updLegend(allMarkers);
  cluster.clearLayers();cluster.addLayers(allMarkers);
  map.setView([-2.5,118],5);
}});
</script>
</body>
</html>"""

# ── Write output ──────────────────────────────────────────────────────────────
os.makedirs("output", exist_ok=True)
out_dated = f"output/knmp_{tgl_file}.html"
out_root  = "knmp.html"

with open(out_dated, "w", encoding="utf-8") as f: f.write(html)
with open(out_root,  "w", encoding="utf-8") as f: f.write(html)

kb = len(html.encode("utf-8")) // 1024
kat_debug = sorted(set(m["kategori"] for m in markers))
print(f"Generated {out_root} ({kb} KB)")
print(f"Generated {out_dated}")
print(f"Total: {total}  Survey: {n_sv}  Ops: {n_op}  Selesai: {selesai}  Kontrak: {fmt_rp(tot_nilai)}")
print(f"Kategori unik: {kat_debug}  Provinsi: {len(prov_all)}")
print(f"Harga wilayah: {len(harga_wilayah)} wilayah dimuat.")
