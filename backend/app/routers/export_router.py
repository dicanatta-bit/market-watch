"""Export router — export Excel dengan filter pulau & provinsi"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from ..database import get_db, SessionLocal
from ..models import KnmpLocation
from ..auth import get_superadmin, get_current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tempfile
from datetime import date

router = APIRouter(tags=["export"])

# ── Mapping provinsi → pulau ──
PROV_WILAYAH = {
    "Sumatera": ["ACEH", "SUMATERA UTARA", "SUMATRA UTARA", "SUMATERA BARAT", "SUMATRA BARAT",
                 "RIAU", "KEPULAUAN RIAU", "JAMBI", "BENGKULU", "SUMATERA SELATAN",
                 "SUMATRA SELATAN", "LAMPUNG", "KEPULAUAN BANGKA BELITUNG", "BANGKA BELITUNG"],
    "Jawa-Bali": ["BANTEN", "DKI JAKARTA", "JAKARTA", "JAWA BARAT", "JAWA TENGAH",
                  "DI YOGYAKARTA", "JAWA TIMUR", "BALI"],
    "Kalimantan": ["KALIMANTAN BARAT", "KALIMANTAN TENGAH", "KALIMANTAN SELATAN",
                   "KALIMANTAN TIMUR", "KALIMANTAN UTARA"],
    "Sulawesi": ["SULAWESI UTARA", "SULAWESI TENGAH", "SULAWESI SELATAN",
                 "SULAWESI TENGGARA", "GORONTALO", "SULAWESI BARAT"],
    "NTT-NTB": ["NUSA TENGGARA BARAT", "NTB", "NUSA TENGGARA TIMUR", "NTT"],
    "Maluku": ["MALUKU", "MALUKU UTARA"],
    "Papua": ["PAPUA", "PAPUA BARAT", "PAPUA SELATAN", "PAPUA TENGAH",
              "PAPUA PEGUNUNGAN", "PAPUA BARAT DAYA", "IRIAN JAYA BARAT"],
}

PULAU_LIST = sorted(PROV_WILAYAH.keys())


@router.get("/export/excel")
def export_excel(
    pulau: str = None,
    provinsi: list[str] = Query(None),
):
    """Export KNMP locations to Excel with optional filtering.

    - `pulau`: filter by island (Sumatera, Jawa-Bali, Kalimantan, Sulawesi, NTT-NTB, Maluku, Papua)
    - `provinsi`: filter by specific provinces (repeatable: ?provinsi=ACEH&provinsi=JAWA+BARAT)
    """
    db = SessionLocal()
    try:
        # Build filter
        prov_filter = set()
        if pulau:
            p = pulau.strip()
            if p in PROV_WILAYAH:
                prov_filter.update(PROV_WILAYAH[p])
        if provinsi:
            for p in provinsi:
                prov_filter.add(p.strip().upper())

        query = db.query(KnmpLocation)
        if prov_filter:
            query = query.filter(KnmpLocation.provinsi.in_(list(prov_filter)))

        locations = query.order_by(KnmpLocation.provinsi, KnmpLocation.kabupaten, KnmpLocation.nama_kampung).all()

        # Build filename
        if pulau:
            filename = f"KNMP_{pulau}_{date.today()}.xlsx"
        elif prov_filter:
            label = "_".join(sorted(prov_filter)[:3])
            filename = f"KNMP_{label}_{date.today()}.xlsx"
        else:
            filename = f"KNMP_Semua_{date.today()}.xlsx"

        # ── Generate Excel ──
        wb = Workbook()
        ws = wb.active
        ws.title = "KNMP Locations"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["ID", "Nama Kampung", "Provinsi", "Kabupaten", "Kecamatan",
                    "Latitude", "Longitude", "Status KNMP", "Nelayan", "Kapal"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row, loc in enumerate(locations, 2):
            data = [
                loc.id_lokasi, loc.nama_kampung, loc.provinsi, loc.kabupaten,
                loc.kecamatan, loc.lat, loc.lon, loc.status_knmp,
                loc.jumlah_nelayan, loc.jumlah_kapal,
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.font = Font(size=9)

        # Auto-width
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 10

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        return FileResponse(tmp.name, filename=filename,
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    finally:
        db.close()
